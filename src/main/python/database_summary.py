# __author__ = 'yma'
# This file is to get data summary from the database, 1/23/2019

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from pandas import DataFrame
import pandas as pd
import pyodbc
import sys
import os

# usage = ("Correct Usage: database_summary.py <project_directory> <scenarioYear> <scenario_id>")

# check if too few/many arguments passed raise error
if len(sys.argv) != 4:
    sys.exit(-1)

output_path = str(sys.argv[1])
sceYear = int(sys.argv[2])
scenario = int(sys.argv[3])

# create settings data frame
settings = {"Parameter": ["Date", "Scenario_id","Year"],
            "Value":     [datetime.now().strftime("%x %X"),scenario,sceYear]}
settings = pd.DataFrame(data=settings)


# set sql server connection to ws
sql_con = pyodbc.connect(driver='{SQL Server}',
                         server='sql2014a8',
                         database='ws',
                         trusted_connection='yes')


### data summary for sensitivty analysis

# 0_scenarioInfor
scenarioInfor = pd.read_sql_query(
    sql="exec [ws].[sst2].[m0_scenario] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 1_modeshare
modeShare = pd.read_sql_query(
    sql="exec [ws].[sst2].[m1_mode_share] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 4_PtripLengthByPurpose
ptripLengthByPurpose = pd.read_sql_query(
    sql="exec [ws].[sst2].[m4_ptrip_distance_purpose] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 5_PtripLengthByMode
ptripDistanceMode = pd.read_sql_query(
    sql="exec [ws].[sst2].[m5_ptrip_distance_mode] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 23_VMT
VMT = pd.read_sql_query(
    sql="exec [ws].[sst2].[m23_vmt_capita] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 6_VHT
VHT = pd.read_sql_query(
    sql="exec [ws].[sst2].[m6_vht_capita] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 7_VHD
VHD = pd.read_sql_query(
    sql="exec [ws].[sst2].[m7_vhd_capita] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 8_TransitBoardingByMode
transitBoardingLinehaulMode = pd.read_sql_query(
    sql="exec [ws].[sst2].[m8_transit_boarding_linehaulmode] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)

# 9_TransitTripsbyMode
transitTripsbyMode = pd.read_sql_query(
    sql="exec [ws].[sst2].[m9_transit_trips_by_mode] @scenario_id = ? ",
    con=sql_con,
    params=[scenario]
)


# output to excel book 'source_sensitivity'
output_file_sens = output_path + "\\analysis\\summary\\source_summary.xlsx"
book = load_workbook(output_file_sens)

with pd.ExcelWriter(output_file_sens, engine='openpyxl') as writer:

    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    settings.to_excel(writer,'Settings',index=False)
    scenarioInfor.to_excel(writer,'0_scenarioInfor', index=False)
    modeShare.to_excel(writer,'1_modeshare', index=False)
    ptripLengthByPurpose.to_excel(writer, '4_PtripLengthByPurpose', index=False)
    ptripDistanceMode.to_excel(writer, '5_PtripLengthByMode', index=False)
    VMT.to_excel(writer, '23_VMT', index=False)
    VHT.to_excel(writer, '6_VHT', index=False)
    VHD.to_excel(writer, '7_VHD', index=False)
    transitBoardingLinehaulMode.to_excel(writer, '8_TransitBoardingByMode', index=False)
    transitTripsbyMode.to_excel(writer, '9_TransitTripsbyMode', index=False)

    writer.save()


### data summary for validation analysis, if scenario_year = 2016 or 2018

if sceYear == 2016 or sceYear == 2018:

    # get hwy flow given scenario_id
    hwyFlow = pd.read_sql_query(
        sql="select * from [ws].[validate2].[FlowDay2016_nonmsa] (?)",
        con=sql_con,
        params=[scenario]
    )

    # get freeway flow by TOD given scenario_id
    fwyFlow = pd.read_sql_query(
        sql="select * from [ws].[validate2].[FlowFreewayTod2016] (?)",
        con=sql_con,
        params=[scenario]
    )

    # get truck flow given scenario_id
    truckFlow = pd.read_sql_query(
        sql="select * from [ws].[validate2].[TruckFlow2016] (?)",
        con=sql_con,
        params=[scenario]
    )

    # get hwy speed given scenario_id
    hwySpeed = pd.read_sql_query(
        sql="select * from [ws].[validate2].[SpeedFreewayTod2016] (?)",
        con=sql_con,
        params=[scenario]
    )

    # get transit general summary given scenario_id
    transitGeneral = pd.read_sql_query(
        sql="select * from [ws].[validate2].[transit2016] (?)",
        con=sql_con,
        params=[scenario]
    )

    # get transit hub summary given scenario_id
    transitHub = pd.read_sql_query(
        sql="select * from [ws].[validate2].[transit2016_Hub] (?)",
        con=sql_con,
        params=[scenario]
    )


    # output to excel book 'source'
    output_file_vald = output_path + "\\analysis\\validation\\source.xlsx"
    book = load_workbook(output_file_vald)

    with pd.ExcelWriter(output_file_vald, engine='openpyxl') as writer:

        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        settings.to_excel(writer,'Settings',index=False)
        hwyFlow.to_excel(writer,'flow_raw', index=False)
        fwyFlow.to_excel(writer,'flow_freeway', index=False)
        truckFlow.to_excel(writer,'truck_flow', index=False)
        hwySpeed.to_excel(writer,'hwy_speed', index=False)
        transitGeneral.to_excel(writer,'transit_general', index=False)
        transitHub.to_excel(writer,'transit_hub', index=False)

    writer.save()

