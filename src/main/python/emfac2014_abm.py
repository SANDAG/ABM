from datetime import datetime
import imp
import pyodbc
import sys
try:
    imp.find_module('_mssql')
except ImportError:
    print 'The current version of Python cannot find the _mssql <http://www.pymssql.org> package.\nPlease ensure _mssql package is properly installed.\nProgram Failed.'
    sys.exit(-1)
'''
try:
    imp.find_module('xlwt')
except ImportError:
    print 'The current version of Python cannot find the xlwt <https://github.com/python-excel/xlwt> package.\nPlease ensure xlwt package is properly installed.\nProgram Failed.'
    sys.exit(-1)
'''
import _mssql
from openpyxl import Workbook
from openpyxl.datavalidation import DataValidation, ValidationType

usage = 'Correct Usage: emfac2014_abm.py <Scenario ID> <Season: Annual | Summer | Winter> <SB 375: On | Off> <Output Path>'

if len(sys.argv) != 5:
    print usage
    sys.exit(-1)
    

scenario_id = int(sys.argv[1])
season = sys.argv[2]
if season not in {'Annual', 'Summer', 'Winter'}:
    print usage
    sys.exit(-1)

sb375 = sys.argv[3]
if sb375 not in {'On', 'Off'}:
    print usage
    sys.exit(-1)
    
saveLocation = sys.argv[4]

#conn = _mssql.connect(server='pele', user='emfac', password='3mf@c', database='abm_sd')
conn = pyodbc.connect(driver='{SQL Server}', server='sql2014a8', database='abm_13_2_3',trusted_connection='yes')
cursor = conn.cursor()

#Get Scenario Year
print 'Retrieving the scenario year'
yearQuery = 'SELECT s.SCENARIO_DESC as sName, s.SCENARIO_YEAR AS fileYear, CASE WHEN s.SCENARIO_YEAR > 2050 THEN 2050 ELSE s.SCENARIO_YEAR END as YR FROM ref.scenario s WHERE SCENARIO_ID = ?'
cursor.execute(yearQuery, scenario_id)
row = cursor.fetchone()
year = row.YR
fileYear = row.fileYear #different from year if 2051 and beyond scenario
scenarioName = row.sName

print 'Creating the EMFAC 2014 User Defined Input Workbook and Worksheets'
book = Workbook()
sheet_settings = book.active
sheet_settings.title = 'Settings'
sheet_scen_vmt = book.create_sheet()
sheet_scen_vmt.title = ('Daily_VMT_By_Veh_Tech')
sheet_scen_speed = book.create_sheet()
sheet_scen_speed.title = 'Hourly_Fraction_Veh_Tech_Speed'

####WRITE REGIONAL SCENARIOS SHEETS####

#Write Column Headers for Regional Scenarios Sheet
print 'Writing header row for Settings worksheet'
dv = DataValidation(validation_type="list", formula1='"Annual,Summer,Winter,January,February,March,April,May,June,July,August,September,October,November,December"', allow_blank=True)
sheet_settings.add_data_validation(dv)

sheet_settings.cell(row = 0, column = 0).value = 'Parameter'
sheet_settings.cell(row = 0, column = 1).value = 'Value'

sheet_settings.cell(row = 1, column = 0).value = 'Date'
sheet_settings.cell(row = 1, column = 1).value = datetime.now().strftime('%x %X')

sheet_settings.cell(row = 2, column = 0).value = 'Season/Month'
sheet_settings.cell(row = 2, column = 1).value = season
dv.add_cell(sheet_settings.cell(row = 2, column = 1))

sheet_settings.cell(row = 3, column = 0).value = 'SB375 Run'
sheet_settings.cell(row = 3, column = 1).value = sb375

sheet_settings.cell(row = 5, column = 1).value = 'You can edit this'
sheet_settings.cell(row = 6, column = 1).value = 'You can not edit this'

#Write Column Headers for Scenario Daily VMT by Veh Tech Sheet
print 'Writing header row for Daily_VMT_By_Veh_Tech worksheet'
sheet_scen_vmt.cell(row = 0, column = 0).value = 'MPO'
sheet_scen_vmt.cell(row = 0, column = 1).value = 'GAI'
sheet_scen_vmt.cell(row = 0, column = 2).value = 'Sub-Area'
sheet_scen_vmt.cell(row = 0, column = 3).value = 'Cal_Year'
sheet_scen_vmt.cell(row = 0, column = 4).value = 'Veh_Tech'
sheet_scen_vmt.cell(row = 0, column = 5).value = 'New Total VMT'

print 'Querying VMT records'
query = 'SELECT * FROM [ws].[emfac].[FN_EMFAC_2014_VMT] (?) order by Veh_Tech'
cursor.execute(query,   scenario_id)
rows = cursor.fetchall()

print 'Writing rows for Daily_VMT_By_Veh_Tech worksheet'
counter = 1
for row in rows:
    sheet_scen_vmt.cell(row = counter, column = 0).value = row.MPO
    sheet_scen_vmt.cell(row = counter, column = 1).value = row.GAI
    sheet_scen_vmt.cell(row = counter, column = 2).value = row[2]
    sheet_scen_vmt.cell(row = counter, column = 3).value = row.Cal_Year
    sheet_scen_vmt.cell(row = counter, column = 4).value = row.Veh_Tech
    sheet_scen_vmt.cell(row = counter, column = 5).value = row[5]
    counter += 1

#Write Column Headers for Scenario Speed Profiles Sheet
print 'Writing header row for Hourly_Fraction_Veh_Tech_Speed worksheet'
sheet_scen_speed.cell(row = 0, column = 0).value = 'MPO'
sheet_scen_speed.cell(row = 0, column = 1).value = 'GAI'
sheet_scen_speed.cell(row = 0, column = 2).value = 'Sub-Area'
sheet_scen_speed.cell(row = 0, column = 3).value = 'Cal_Year'
sheet_scen_speed.cell(row = 0, column = 4).value = 'Veh_Tech'
sheet_scen_speed.cell(row = 0, column = 5).value = 'Hour'
sheet_scen_speed.cell(row = 0, column = 6).value = '5mph'
sheet_scen_speed.cell(row = 0, column = 7).value = '10mph'
sheet_scen_speed.cell(row = 0, column = 8).value = '15mph'
sheet_scen_speed.cell(row = 0, column = 9).value = '20mph'
sheet_scen_speed.cell(row = 0, column = 10).value = '25mph'
sheet_scen_speed.cell(row = 0, column = 11).value = '30mph'
sheet_scen_speed.cell(row = 0, column = 12).value = '35mph'
sheet_scen_speed.cell(row = 0, column = 13).value = '40mph'
sheet_scen_speed.cell(row = 0, column = 14).value = '45mph'
sheet_scen_speed.cell(row = 0, column = 15).value = '50mph'
sheet_scen_speed.cell(row = 0, column = 16).value = '55mph'
sheet_scen_speed.cell(row = 0, column = 17).value = '60mph'
sheet_scen_speed.cell(row = 0, column = 18).value = '65mph'
sheet_scen_speed.cell(row = 0, column = 19).value = '70mph'
sheet_scen_speed.cell(row = 0, column = 20).value = '75mph'
sheet_scen_speed.cell(row = 0, column = 21).value = '80mph'
sheet_scen_speed.cell(row = 0, column = 22).value = '85mph'
sheet_scen_speed.cell(row = 0, column = 23).value = '90mph'

counter = 1

query = 'SELECT * FROM ws.emfac.FN_EMFAC_2014_VMT_SPEED (?) ORDER BY Veh_Tech, Hour'
print 'Querying VMT Speed records'
cursor.execute(query, scenario_id)
rows = cursor.fetchall()
    
print 'Writing rows for Hourly_Fraction_Veh_Tech_Speed worksheet.'
for row in rows:
    sheet_scen_speed.cell(row = counter, column = 0).value = row.MPO
    sheet_scen_speed.cell(row = counter, column = 1).value = row.GAI
    sheet_scen_speed.cell(row = counter, column = 2).value = row[2]
    sheet_scen_speed.cell(row = counter, column = 3).value = row.Cal_Year
    sheet_scen_speed.cell(row = counter, column = 4).value = row.Veh_Tech
    sheet_scen_speed.cell(row = counter, column = 5).value = row.Hour
    sheet_scen_speed.cell(row = counter, column = 6).value = row[6]
    sheet_scen_speed.cell(row = counter, column = 7).value = row[7]
    sheet_scen_speed.cell(row = counter, column = 8).value = row[8]
    sheet_scen_speed.cell(row = counter, column = 9).value = row[9]
    sheet_scen_speed.cell(row = counter, column = 10).value = row[10]
    sheet_scen_speed.cell(row = counter, column = 11).value = row[11]
    sheet_scen_speed.cell(row = counter, column = 12).value = row[12]
    sheet_scen_speed.cell(row = counter, column = 13).value = row[13]
    sheet_scen_speed.cell(row = counter, column = 14).value = row[14]
    sheet_scen_speed.cell(row = counter, column = 15).value = row[15]
    sheet_scen_speed.cell(row = counter, column = 16).value = row[16]
    sheet_scen_speed.cell(row = counter, column = 17).value = row[17]
    sheet_scen_speed.cell(row = counter, column = 18).value = row[18]
    sheet_scen_speed.cell(row = counter, column = 19).value = row[19]
    sheet_scen_speed.cell(row = counter, column = 20).value = row[20]
    sheet_scen_speed.cell(row = counter, column = 21).value = row[21]
    sheet_scen_speed.cell(row = counter, column = 22).value = row[22]
    sheet_scen_speed.cell(row = counter, column = 23).value = row[23]
        
    counter += 1

outLocation = saveLocation + '\EMFAC2014-SANDAG-'+str(scenarioName)+'-'+season+'-'+str(fileYear)
if sb375 == 'On':
    outLocation = outLocation + '-sb375'
outLocation = outLocation + '.xlsx'
print 'Saving worksheet to ' + outLocation
book.save(outLocation)
conn.close()
